"""
SG Datalytics — Bank of Ghana Statistical Bulletin Parser
=========================================================
HOW TO USE:
  1. Go to https://www.bog.gov.gh/publications/statistical-bulletin/
  2. Download the latest Excel bulletin (.xlsx)
  3. Place it in the same folder as this script
  4. Run: python bog_bulletin_parser.py --file "your_file.xlsx"
     Or:  python bog_bulletin_parser.py  (auto-detects latest .xlsx in folder)

OUTPUT: Clean CSV files in ./output/bog/ folder, ready for SG Datalytics.
"""

import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime
import os
import sys
import glob
import argparse

OUTPUT_DIR = "output/bog"
os.makedirs(OUTPUT_DIR, exist_ok=True)
TODAY = datetime.today().strftime("%Y-%m-%d")


# ─────────────────────────────────────────────────────────────
# UTILITY: Read a sheet into a raw list of rows
# ─────────────────────────────────────────────────────────────
def read_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


# ─────────────────────────────────────────────────────────────
# PARSER 1: Foreign Exchange Rates (Sheet 25)
# Monthly USD/GHS, GBP/GHS, EUR/GHS, CHF/GHS, JPY/GHS
# ─────────────────────────────────────────────────────────────
def parse_fx_rates(wb):
    print("  → Parsing FX Rates (Sheet 25)...")
    rows = read_sheet(wb, "25")

    records = []
    for row in rows[4:]:  # skip header rows
        if row[0] is None:
            continue
        try:
            date = pd.to_datetime(row[0])
        except Exception:
            continue

        record = {
            "Date": date.strftime("%Y-%m"),
            "Year": date.year,
            "Month": date.strftime("%B"),
            "Source": "Bank of Ghana Statistical Bulletin",
            "Scraped Date": TODAY,
        }

        # Columns: USD end, USD avg, GBP end, GBP avg, EUR end, EUR avg,
        #          CHF end, CHF avg, JPY end, JPY avg
        currency_map = {
            "USD_GHS_End_Period": 1,
            "USD_GHS_Period_Average": 2,
            "GBP_GHS_End_Period": 3,
            "GBP_GHS_Period_Average": 4,
            "EUR_GHS_End_Period": 5,
            "EUR_GHS_Period_Average": 6,
            "CHF_GHS_End_Period": 7,
            "CHF_GHS_Period_Average": 8,
            "JPY_GHS_End_Period": 9,
            "JPY_GHS_Period_Average": 10,
        }

        for col_name, col_idx in currency_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            try:
                record[col_name] = round(float(val), 6) if val is not None else None
            except (TypeError, ValueError):
                record[col_name] = None

        records.append(record)

    df = pd.DataFrame(records)
    df = df.dropna(subset=["USD_GHS_End_Period"])
    df = df.sort_values("Date", ascending=False).reset_index(drop=True)

    filepath = f"{OUTPUT_DIR}/bog_fx_rates_{TODAY}.csv"
    df.to_csv(filepath, index=False)
    print(f"     ✓ {len(df)} monthly FX records → {filepath}")
    return df


# ─────────────────────────────────────────────────────────────
# PARSER 2: CPI & Inflation Rates (Sheet 27)
# ─────────────────────────────────────────────────────────────
def parse_cpi_inflation(wb):
    print("  → Parsing CPI & Inflation (Sheet 27)...")
    rows = read_sheet(wb, "27")

    # Find the "YEAR-ON-YEAR INFLATION RATE" section
    # and extract Headline Inflation row
    # Structure: row[0]=indicator name, then columns are months across years

    # First, build month-year headers from rows 1-2
    # Row 1: years (2013, 2014, ...)
    # Row 2: months (Jan, Feb, ..., Dec)

    year_row = rows[1]
    month_row = rows[2]

    # Build column date labels
    col_dates = []
    current_year = None
    for i in range(1, len(year_row)):
        if year_row[i] is not None:
            try:
                current_year = int(year_row[i])
            except (ValueError, TypeError):
                pass
        month = month_row[i] if i < len(month_row) else None
        if current_year and month:
            col_dates.append(f"{current_year}-{str(month)[:3]}")
        else:
            col_dates.append(None)

    # Extract rows we care about
    target_indicators = {
        "OVERALL INDEX": "CPI Overall Index",
        "Food and Non-Alcoholic Be": "CPI Food & Non-Alcoholic Beverages",
        "Non-Food Group": "CPI Non-Food",
        "Headline Inflation": "Headline Inflation (%)",
        "Core 1: Inflation excl En": "Core Inflation excl Energy (%)",
    }

    records_by_date = {}
    in_inflation_section = False

    for row in rows[3:]:
        indicator = str(row[0]).strip() if row[0] else ""

        if "YEAR-ON-YEAR" in indicator.upper():
            in_inflation_section = True

        matched_key = None
        for search_str, clean_name in target_indicators.items():
            if indicator.startswith(search_str):
                matched_key = clean_name
                break

        if matched_key:
            for i, date_label in enumerate(col_dates):
                if date_label is None:
                    continue
                val = row[i + 1] if (i + 1) < len(row) else None
                try:
                    val = round(float(val), 4) if val is not None else None
                except (TypeError, ValueError):
                    val = None

                if date_label not in records_by_date:
                    records_by_date[date_label] = {
                        "Date": date_label,
                        "Source": "Bank of Ghana Statistical Bulletin",
                        "Scraped Date": TODAY,
                    }
                records_by_date[date_label][matched_key] = val

    df = pd.DataFrame(list(records_by_date.values()))
    if df.empty:
        print("     ✗ No CPI data extracted")
        return None

    # Parse date for sorting
    df["_sort"] = pd.to_datetime(df["Date"], format="%Y-%b", errors="coerce")
    df = df.sort_values("_sort", ascending=False).drop(columns=["_sort"]).reset_index(drop=True)
    df = df.dropna(subset=["Headline Inflation (%)"])

    filepath = f"{OUTPUT_DIR}/bog_cpi_inflation_{TODAY}.csv"
    df.to_csv(filepath, index=False)
    print(f"     ✓ {len(df)} monthly CPI/inflation records → {filepath}")
    return df


# ─────────────────────────────────────────────────────────────
# PARSER 3: Selected Economic Indicators Summary
# ─────────────────────────────────────────────────────────────
def parse_selected_indicators(wb):
    print("  → Parsing Selected Economic Indicators...")
    rows = read_sheet(wb, "Selected Indicators")

    # This sheet has a complex multi-year, multi-month layout
    # Row 0: title, Row 1: years, Row 2: months (Jan-Dec)

    year_row = rows[1]
    month_row = rows[2]

    col_dates = []
    current_year = None
    for i in range(1, len(year_row)):
        yr = year_row[i]
        mo = month_row[i] if i < len(month_row) else None
        if yr is not None:
            try:
                current_year = int(yr)
            except (TypeError, ValueError):
                pass
        if current_year and mo:
            col_dates.append(f"{current_year}-{str(mo)[:3]}")
        else:
            col_dates.append(None)

    # Key indicators to extract
    targets = {
        "Headline Inflation": "Headline Inflation (%)",
        "Core 1: Inflation excl En": "Core Inflation (%)",
        "Private Sector Credit": "Private Sector Credit Growth (%)",
        "Reserve Money": "Reserve Money Growth (%)",
        "Broad Money (M2)": "Broad Money M2 Growth (%)",
        "Gross International Reser": "Gross International Reserves (USD M)",
        "Policy Rate": "Monetary Policy Rate (%)",
        "91-day T-Bill Rate": "91-Day T-Bill Rate (%)",
    }

    records_by_date = {}
    for row in rows[3:]:
        indicator = str(row[0]).strip() if row[0] else ""
        for search_str, clean_name in targets.items():
            if indicator.startswith(search_str):
                for i, date_label in enumerate(col_dates):
                    if date_label is None:
                        continue
                    val = row[i + 1] if (i + 1) < len(row) else None
                    try:
                        val = round(float(val), 4) if val is not None else None
                    except (TypeError, ValueError):
                        val = None
                    if date_label not in records_by_date:
                        records_by_date[date_label] = {
                            "Date": date_label,
                            "Source": "Bank of Ghana Statistical Bulletin",
                            "Scraped Date": TODAY,
                        }
                    records_by_date[date_label][clean_name] = val

    df = pd.DataFrame(list(records_by_date.values()))
    if df.empty:
        print("     ✗ No indicators extracted")
        return None

    df["_sort"] = pd.to_datetime(df["Date"], format="%Y-%b", errors="coerce")
    df = df.sort_values("_sort", ascending=False).drop(columns=["_sort"]).reset_index(drop=True)

    filepath = f"{OUTPUT_DIR}/bog_key_indicators_{TODAY}.csv"
    df.to_csv(filepath, index=False)
    print(f"     ✓ {len(df)} monthly key indicator records → {filepath}")
    return df


# ─────────────────────────────────────────────────────────────
# PARSER 4: GDP Data (Sheets 28 & 29)
# ─────────────────────────────────────────────────────────────
def parse_gdp(wb):
    print("  → Parsing GDP Data (Sheet 28 & 29)...")

    all_records = []
    for sheet_name in ["28", "29"]:
        if sheet_name not in wb.sheetnames:
            continue
        rows = read_sheet(wb, sheet_name)

        for row in rows[4:]:
            if not row[0]:
                continue
            quarter = str(row[0]).strip()
            if not any(q in quarter for q in ["Q1", "Q2", "Q3", "Q4", "2010", "2011",
                                               "2012", "2013", "2014", "2015", "2016",
                                               "2017", "2018", "2019", "2020", "2021",
                                               "2022", "2023", "2024", "2025"]):
                continue

            record = {
                "Period": quarter,
                "Source": "Bank of Ghana Statistical Bulletin",
                "Scraped Date": TODAY,
                "Sheet": f"Table {sheet_name}",
            }

            # Try to extract numeric values from columns 1-5
            labels = ["GDP Total", "Agriculture", "Industry", "Services", "GDP Growth (%)"]
            for j, label in enumerate(labels):
                val = row[j + 1] if (j + 1) < len(row) else None
                try:
                    record[label] = round(float(val), 4) if val is not None else None
                except (TypeError, ValueError):
                    record[label] = None

            all_records.append(record)

    if not all_records:
        print("     ✗ No GDP data extracted")
        return None

    df = pd.DataFrame(all_records)
    filepath = f"{OUTPUT_DIR}/bog_gdp_{TODAY}.csv"
    df.to_csv(filepath, index=False)
    print(f"     ✓ {len(df)} quarterly GDP records → {filepath}")
    return df


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Parse Bank of Ghana Statistical Bulletin")
    parser.add_argument("--file", help="Path to the .xlsx bulletin file", default=None)
    args = parser.parse_args()

    # Auto-detect file if not specified
    if args.file:
        xlsx_path = args.file
    else:
        xlsx_files = glob.glob("bog_bulletins/*.xlsx") + glob.glob("*.xlsx")
        if not xlsx_files:
            print("❌ No .xlsx file found. Please download from:")
            print("   https://www.bog.gov.gh/publications/statistical-bulletin/")
            print("   Then run: python bog_bulletin_parser.py --file your_file.xlsx")
            sys.exit(1)
        xlsx_path = sorted(xlsx_files)[-1]  # Use most recently modified

    print("=" * 60)
    print("  SG DATALYTICS — Bank of Ghana Bulletin Parser")
    print(f"  File: {os.path.basename(xlsx_path)}")
    print(f"  Date: {datetime.today().strftime('%d %B %Y')}")
    print("=" * 60)

    print(f"\n📂 Loading workbook...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    print(f"   ✓ {len(wb.sheetnames)} sheets found")

    print("\n🔄 Parsing datasets...")
    results = {
        "FX Rates": parse_fx_rates(wb),
        "CPI & Inflation": parse_cpi_inflation(wb),
        "Key Indicators": parse_selected_indicators(wb),
        "GDP": parse_gdp(wb),
    }

    print("\n" + "=" * 60)
    print("  ✅ SUMMARY")
    print("=" * 60)
    total = 0
    for name, df in results.items():
        if df is not None and not df.empty:
            print(f"  ✓ {name:<25} {len(df):>4} rows → output/bog/")
            total += len(df)
        else:
            print(f"  ✗ {name:<25} Failed or empty")

    print(f"\n  Total records: {total}")
    print(f"  Output folder: ./{OUTPUT_DIR}/")
    print("=" * 60)

    # Preview FX rates
    fx = results.get("FX Rates")
    if fx is not None and not fx.empty:
        print("\n  📊 Latest FX Rates (USD/GHS):")
        print(fx[["Date", "USD_GHS_End_Period", "USD_GHS_Period_Average",
                   "GBP_GHS_End_Period", "EUR_GHS_End_Period"]].head(6).to_string(index=False))

    # Preview inflation
    cpi = results.get("CPI & Inflation")
    if cpi is not None and not cpi.empty:
        print("\n  📊 Latest Inflation Rates:")
        cols = [c for c in ["Date", "Headline Inflation (%)", "Core Inflation excl Energy (%)"] if c in cpi.columns]
        print(cpi[cols].head(6).to_string(index=False))


if __name__ == "__main__":
    main()
