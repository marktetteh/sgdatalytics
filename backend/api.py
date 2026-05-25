"""
SG Datalytics — SGMPI REST API v2
Connects to 6 Neon databases serving real Ghana market price data.

Run locally : python3 api.py
Run on Railway: gunicorn -w 2 -b 0.0.0.0:$PORT api:app
"""
import os, csv, io, hmac, hashlib, secrets
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta, timezone
from flask import Flask, jsonify, request, Response, stream_with_context
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import psycopg2, psycopg2.extras

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

app = Flask(__name__)
CORS(app)

# ── RATE LIMITER ─────────────────────────────────────────────
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per hour"],
    storage_uri="memory://",
)

# ── DB CONNECTIONS ───────────────────────────────────────────
DB = {
    'market_prices':  os.getenv('NEON_MARKET_PRICES'),
    'accommodation':  os.getenv('NEON_ACCOMMODATION'),
    'property':       os.getenv('NEON_PROPERTY'),
    'economic':       os.getenv('NEON_ECONOMIC'),
    'commodities':    os.getenv('NEON_COMMODITIES'),
    'financials':     os.getenv('NEON_FINANCIALS'),
}

def get_conn(db_key):
    url = DB.get(db_key)
    if not url:
        raise Exception(f'No connection string for {db_key}')
    return psycopg2.connect(url, sslmode='require')

def query(db_key, sql, params=None, one=False):
    conn = get_conn(db_key)
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(sql, params or [])
    result = cur.fetchone() if one else cur.fetchall()
    cur.close(); conn.close()
    return result

# ── TOKEN STORE (Neon-backed) ─────────────────────────────────
# Tokens are persisted in the economic Neon DB so they survive Railway redeploys.
# Table is auto-created on first use.

def _get_token_conn():
    url = os.getenv('NEON_ECONOMIC', '')
    if not url:
        raise Exception('NEON_ECONOMIC not configured — cannot store tokens')
    return psycopg2.connect(url, sslmode='require')

def _ensure_token_table():
    """Create download_tokens table if it doesn't exist."""
    try:
        conn = _get_token_conn()
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS download_tokens (
                        token       TEXT PRIMARY KEY,
                        email       TEXT NOT NULL,
                        sector      TEXT NOT NULL,
                        created_at  TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        expires_at  TIMESTAMPTZ NOT NULL,
                        used        BOOLEAN NOT NULL DEFAULT FALSE
                    )
                """)
        conn.close()
    except Exception as e:
        print(f"[TOKEN] Warning: could not ensure token table: {e}")

# Create table on startup
_ensure_token_table()

# ── SECTOR CONFIG ────────────────────────────────────────────
SECTOR_LABELS = {
    'market_prices': 'Market Prices',
    'accommodation': 'Real Estate & Accommodation',
    'economic':      'Economic, Financial & Agricultural Data',
    'bundle':        'Ghana Complete Data Bundle',
}

# Bundle now only needs 3 download links — all 8 tables are covered
BUNDLE_SECTORS = ['market_prices', 'accommodation', 'economic']

# Exact product key → sector mapping (used with Paystack metadata.product)
PRODUCT_SECTOR_MAP = {
    'market_prices': 'market_prices',
    'accommodation': 'accommodation',
    'economic':      'economic',
    'bundle':        'bundle',
}

# Each entry: (db_key, sql, table_label)
# table_label is written as a section header in the CSV so analysts know where each table starts
SECTOR_QUERIES = {
    # ── Consumer Market Prices ───────────────────────────────────
    # Aggregated like Numbeo: one row per product per city per week
    # Outliers filtered: price_ghs between GHS 1 and GHS 200,000
    'market_prices': [
        # Table 1 — Category summary: one row per category per city per week
        ('market_prices', """
        SELECT
            week_number,
            year,
            product_category,
            COALESCE(NULLIF(location, ''), 'Ghana')               AS city,
            COUNT(*)                                              AS listing_count,
            ROUND(AVG(price_ghs)::numeric,    0)                  AS avg_price_ghs,
            MIN(price_ghs)                                        AS min_price_ghs,
            MAX(price_ghs)                                        AS max_price_ghs,
            ROUND(PERCENTILE_CONT(0.5) WITHIN GROUP
                  (ORDER BY price_ghs)::numeric, 0)               AS median_price_ghs
        FROM market_prices
        WHERE price_ghs > 1
          AND price_ghs < 200000
        GROUP BY week_number, year,
                 product_category,
                 COALESCE(NULLIF(location, ''), 'Ghana')
        ORDER BY year DESC, week_number DESC, product_category, city
        """, 'market_prices_category_summary'),

        # Table 2 — Product detail: one row per product name per city per week
        ('market_prices', """
        SELECT
            week_number,
            year,
            product_category,
            TRIM(COALESCE(NULLIF(normalized_name, ''), search_label)) AS normalized_name,
            COALESCE(NULLIF(location, ''), 'Ghana')               AS city,
            COUNT(*)                                              AS listing_count,
            ROUND(AVG(price_ghs)::numeric,    0)                  AS avg_price_ghs,
            MIN(price_ghs)                                        AS min_price_ghs,
            MAX(price_ghs)                                        AS max_price_ghs,
            ROUND(PERCENTILE_CONT(0.5) WITHIN GROUP
                  (ORDER BY price_ghs)::numeric, 0)               AS median_price_ghs
        FROM market_prices
        WHERE price_ghs > 1
          AND price_ghs < 200000
          AND (normalized_name IS NOT NULL AND normalized_name <> ''
               OR search_label IS NOT NULL)
        GROUP BY week_number, year,
                 product_category,
                 TRIM(COALESCE(NULLIF(normalized_name, ''), search_label)),
                 COALESCE(NULLIF(location, ''), 'Ghana')
        ORDER BY year DESC, week_number DESC, product_category, normalized_name
        """, 'market_prices_product_detail'),
    ],

    # ── Real Estate & Accommodation ──────────────────────────────
    'accommodation': [
        ('accommodation', """
        SELECT
            week_number,
            year,
            'Hotel'                                        AS product_category,
            city || ' — ' || COALESCE(star_rating::text, '?') || '-star hotel'
                                                           AS normalized_name,
            city,
            COALESCE(star_rating::text, 'Unknown')         AS star_rating,
            COUNT(*)                                        AS hotel_count,
            ROUND(AVG(price_per_night_usd)::numeric, 2)    AS avg_price_per_night_usd,
            MIN(price_per_night_usd)                        AS min_price_per_night_usd,
            MAX(price_per_night_usd)                        AS max_price_per_night_usd,
            ROUND(AVG(review_score)::numeric, 1)            AS avg_review_score
        FROM hotel_prices
        WHERE price_per_night_usd > 0
        GROUP BY week_number, year, city, star_rating
        ORDER BY year DESC, week_number DESC, city, star_rating
        """, 'hotel_prices_aggregated'),

        ('accommodation', """
        SELECT
            week_number,
            year,
            'Airbnb'                                       AS product_category,
            city || ' — ' || room_type                     AS normalized_name,
            city,
            room_type,
            COUNT(*)                                        AS listing_count,
            ROUND(AVG(price_ghs)::numeric, 0)              AS avg_price_ghs,
            MIN(price_ghs)                                  AS min_price_ghs,
            MAX(price_ghs)                                  AS max_price_ghs,
            ROUND(AVG(rating)::numeric, 1)                  AS avg_rating
        FROM airbnb_prices
        WHERE price_ghs > 0
        GROUP BY week_number, year, city, room_type
        ORDER BY year DESC, week_number DESC, city, room_type
        """, 'airbnb_prices_aggregated'),

        ('property', """
        SELECT
            week_number,
            year,
            'Property'                                     AS product_category,
            city || ' — ' || property_type
              || CASE WHEN bedrooms IS NOT NULL
                      THEN ' ' || bedrooms || 'BR' ELSE '' END
                                                           AS normalized_name,
            property_type,
            listing_type,
            city,
            neighborhood,
            bedrooms,
            COUNT(*)                                        AS listing_count,
            ROUND(AVG(price_ghs)::numeric, 0)              AS avg_price_ghs,
            MIN(price_ghs)                                  AS min_price_ghs,
            MAX(price_ghs)                                  AS max_price_ghs
        FROM property_prices
        WHERE price_ghs > 0
        GROUP BY week_number, year,
                 property_type, listing_type, city, neighborhood, bedrooms
        ORDER BY year DESC, week_number DESC, city, property_type, bedrooms
        """, 'property_prices_aggregated'),
    ],

    # ── Economic, Financial & Agricultural ───────────────────────
    'economic': [
        ('economic', """
        SELECT
            collected_date,
            year,
            month,
            sector                 AS product_category,
            indicator_name         AS normalized_name,
            indicator_code,
            value,
            unit,
            source
        FROM economic_indicators
        ORDER BY collected_date DESC, sector, indicator_name
        """, 'economic_indicators'),

        ('economic', """
        SELECT
            collected_date,
            'Foreign Exchange'     AS product_category,
            currency_pair          AS normalized_name,
            currency_pair,
            rate_ghs,
            source
        FROM exchange_rates
        ORDER BY collected_date DESC, currency_pair
        """, 'exchange_rates'),

        ('financials', """
        SELECT
            collected_date,
            week_number,
            year,
            'Financial Markets'    AS product_category,
            index_name             AS normalized_name,
            index_name,
            value,
            change_points,
            change_pct,
            source
        FROM gse_indices
        ORDER BY collected_date DESC, index_name
        """, 'gse_indices'),

        ('financials', """
        SELECT
            collected_date,
            week_number,
            year,
            'Financial Markets'    AS product_category,
            company_name || ' (' || symbol || ')'
                                   AS normalized_name,
            symbol,
            company_name,
            opening_price_ghs,
            closing_price_ghs,
            change_ghs,
            change_pct,
            volume,
            year_high,
            year_low,
            source
        FROM stock_prices
        ORDER BY collected_date DESC, symbol
        """, 'stock_prices'),

        ('commodities', """
        SELECT
            collected_date,
            week_number,
            year,
            'Agricultural Commodities'   AS product_category,
            commodity_name || ' — ' || REPLACE(market, 'Esoko Marketplace', 'Accra Retail Prices')
                                         AS normalized_name,
            commodity_name,
            REPLACE(market, 'Esoko Marketplace', 'Accra Retail Prices') AS market,
            region,
            price_ghs,
            unit,
            source
        FROM commodity_prices
        ORDER BY collected_date DESC, commodity_name, market
        """, 'commodity_prices'),

        ('commodities', """
        SELECT
            collected_date,
            week_number,
            year,
            'Energy & Fuel'        AS product_category,
            fuel_type              AS normalized_name,
            fuel_type,
            price_ghs_per_litre,
            currency,
            source
        FROM fuel_prices
        ORDER BY collected_date DESC, fuel_type
        """, 'fuel_prices'),
    ],
}

# Maps Paystack plan name keywords → sector codes
PLAN_SECTOR_MAP = {
    'market':        'market_prices',
    'property':      'property',
    'accommodation': 'accommodation',
    'hotel':         'accommodation',
    'airbnb':        'accommodation',
    'economic':      'economic',
    'commodit':      'economic',
    'agricultur':    'economic',
    'financial':     'economic',
    'fuel':          'commodities',
    'financial':     'financials',
    'stock':         'financials',
    'gse':           'financials',
}

def resolve_sector(plan_name):
    """Map a Paystack plan name to a sector code using keyword matching."""
    name_lower = (plan_name or '').lower()
    for keyword, sector in PLAN_SECTOR_MAP.items():
        if keyword in name_lower:
            return sector
    return 'market_prices'  # safe default

# ═══════════════════════════════════════════════════════════════
# PART 2 — TOKEN GENERATION
# ═══════════════════════════════════════════════════════════════
def generate_download_token(email, sector, expires_hours=24):
    """
    Generate a secure one-time download token persisted in Neon.
    Survives Railway redeploys — tokens valid for 24 hours.
    """
    token      = secrets.token_urlsafe(32)
    now        = datetime.now(timezone.utc)
    expires_at = now + timedelta(hours=expires_hours)

    try:
        conn = _get_token_conn()
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO download_tokens (token, email, sector, created_at, expires_at, used)
                    VALUES (%s, %s, %s, %s, %s, FALSE)
                """, (token, email, sector, now, expires_at))
        conn.close()
    except Exception as e:
        print(f"[TOKEN] DB write failed: {e} — token will not persist across redeploys")

    download_url = f"https://sgdatalytics-production.up.railway.app/api/download?token={token}"
    print(f"[TOKEN] Generated for {email} | sector={sector} | expires={expires_at.isoformat()}")
    return download_url

# ═══════════════════════════════════════════════════════════════
# PART 3 — EMAIL SENDING
# ═══════════════════════════════════════════════════════════════
def send_download_email(email, download_url, sector):
    """
    Send a professional HTML email with the one-time download link.
    Uses Resend HTTP API (RESEND_API_KEY env var) — no SMTP, Railway-safe.
    """
    import requests as req
    api_key = os.getenv('RESEND_API_KEY', '').strip()
    if not api_key:
        raise Exception('RESEND_API_KEY not configured')

    sector_label = SECTOR_LABELS.get(sector, sector.replace('_', ' ').title())

    html = f"""<!DOCTYPE html>
<html>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0">
  <tr><td align="center" style="padding:40px 20px;">
    <table width="600" style="background:#ffffff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.08);">
      <tr><td style="background:#00957a;padding:32px;text-align:center;">
        <h1 style="color:#ffffff;margin:0;font-size:26px;letter-spacing:-0.5px;">SG Datalytics</h1>
        <p style="color:#d0f0e8;margin:8px 0 0;font-size:14px;">Ghana Market Intelligence Platform</p>
      </td></tr>
      <tr><td style="padding:40px;">
        <h2 style="color:#1a1a1a;margin:0 0 12px;font-size:20px;">Your data is ready ✓</h2>
        <p style="color:#555;line-height:1.7;margin:0 0 24px;">
          Thank you for your purchase. Your <strong>{sector_label}</strong> dataset
          has been prepared and is ready to download as a CSV file.
        </p>
        <div style="text-align:center;margin:32px 0;">
          <a href="{download_url}"
             style="background:#00957a;color:#ffffff;padding:16px 48px;border-radius:6px;
                    text-decoration:none;font-size:16px;font-weight:bold;display:inline-block;">
            ⬇ Download Your Data
          </a>
        </div>
        <div style="background:#fff8e1;border-left:4px solid #f59e0b;padding:16px 20px;
                    border-radius:4px;margin:24px 0;">
          <p style="margin:0;color:#92400e;font-size:14px;line-height:1.6;">
            ⚠️ <strong>Important:</strong> This link expires in <strong>24 hours</strong>
            and can only be used <strong>once</strong>. Please download your file immediately.
          </p>
        </div>
        <p style="color:#777;font-size:13px;line-height:1.6;">
          If the button doesn't work, copy and paste this link into your browser:<br>
          <a href="{download_url}" style="color:#00957a;word-break:break-all;">{download_url}</a>
        </p>
        <hr style="border:none;border-top:1px solid #eee;margin:28px 0;">
        <p style="color:#999;font-size:12px;margin:0;">
          Questions? <a href="mailto:data@sgdatalytics.org" style="color:#00957a;">data@sgdatalytics.org</a>
        </p>
      </td></tr>
      <tr><td style="background:#f9f9f9;padding:20px;text-align:center;border-top:1px solid #eee;">
        <p style="color:#bbb;font-size:12px;margin:0;">
          &copy; 2026 SG Datalytics &nbsp;|&nbsp;
          <a href="https://sgdatalytics.org" style="color:#00957a;text-decoration:none;">sgdatalytics.org</a>
        </p>
      </td></tr>
    </table>
  </td></tr>
</table>
</body>
</html>"""

    resp = req.post(
        'https://api.resend.com/emails',
        headers={'Authorization': f'Bearer {api_key}', 'Content-Type': 'application/json'},
        json={
            'from':    'SG Datalytics <data@sgdatalytics.org>',
            'to':      [email],
            'subject': f'Your SG Datalytics Data is Ready — {sector_label}',
            'html':    html,
        },
        timeout=15,
    )
    if resp.status_code not in (200, 201):
        raise Exception(f'Resend API error {resp.status_code}: {resp.text[:200]}')

    print(f"[EMAIL] Sent via Resend to {email} | sector={sector}")


def send_bundle_email(email, sectors):
    """
    Send one email containing separate download links for each sector in the bundle.
    Called when a customer purchases the 'Ghana Complete Data Bundle'.
    """
    import requests as req
    api_key = os.getenv('RESEND_API_KEY', '').strip()
    if not api_key:
        raise Exception('RESEND_API_KEY not configured')

    # Generate a token for each sector
    links_html = ''
    for sector in sectors:
        url   = generate_download_token(email, sector)
        label = SECTOR_LABELS.get(sector, sector.replace('_', ' ').title())
        links_html += f"""
        <tr>
          <td style="padding:14px 0;border-bottom:1px solid #eee;">
            <strong style="color:#1a1a1a;">{label}</strong><br>
            <a href="{url}"
               style="display:inline-block;margin-top:8px;background:#00957a;color:#fff;
                      padding:10px 28px;border-radius:5px;text-decoration:none;font-size:14px;font-weight:bold;">
              ⬇ Download
            </a>
          </td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0">
  <tr><td align="center" style="padding:40px 20px;">
    <table width="620" style="background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.08);">
      <tr><td style="background:#00957a;padding:32px;text-align:center;">
        <h1 style="color:#fff;margin:0;font-size:26px;">SG Datalytics</h1>
        <p style="color:#d0f0e8;margin:8px 0 0;font-size:14px;">Ghana Market Intelligence Platform</p>
      </td></tr>
      <tr><td style="padding:40px;">
        <h2 style="color:#1a1a1a;margin:0 0 12px;font-size:20px;">Your Complete Bundle is Ready ✓</h2>
        <p style="color:#555;line-height:1.7;margin:0 0 24px;">
          Thank you for purchasing the <strong>Ghana Complete Data Bundle</strong>.
          Below are your individual download links — one per dataset.
        </p>
        <div style="background:#fff8e1;border-left:4px solid #f59e0b;padding:14px 18px;
                    border-radius:4px;margin:0 0 24px;">
          <p style="margin:0;color:#92400e;font-size:13px;line-height:1.6;">
            ⚠️ Each link is <strong>one-time use</strong> and expires in <strong>24 hours</strong>.
            Download all files now.
          </p>
        </div>
        <table width="100%" cellpadding="0" cellspacing="0">
          {links_html}
        </table>
        <hr style="border:none;border-top:1px solid #eee;margin:28px 0;">
        <p style="color:#999;font-size:12px;margin:0;">
          Questions? <a href="mailto:data@sgdatalytics.org" style="color:#00957a;">data@sgdatalytics.org</a>
        </p>
      </td></tr>
      <tr><td style="background:#f9f9f9;padding:20px;text-align:center;border-top:1px solid #eee;">
        <p style="color:#bbb;font-size:12px;margin:0;">
          &copy; 2026 SG Datalytics &nbsp;|&nbsp;
          <a href="https://sgdatalytics.org" style="color:#00957a;text-decoration:none;">sgdatalytics.org</a>
        </p>
      </td></tr>
    </table>
  </td></tr>
</table>
</body>
</html>"""

    resp = req.post(
        'https://api.resend.com/emails',
        headers={'Authorization': f'Bearer {api_key}', 'Content-Type': 'application/json'},
        json={
            'from':    'SG Datalytics <data@sgdatalytics.org>',
            'to':      [email],
            'subject': 'Your SG Datalytics Complete Bundle — 4 Download Links Inside',
            'html':    html,
        },
        timeout=15,
    )
    if resp.status_code not in (200, 201):
        raise Exception(f'Resend API error {resp.status_code}: {resp.text[:200]}')

    print(f"[EMAIL] Bundle sent via Resend to {email} | {len(sectors)} sectors")


# ═══════════════════════════════════════════════════════════════
# EXISTING ROUTES (unchanged)
# ═══════════════════════════════════════════════════════════════

@app.route('/')
def root():
    return jsonify({
        "name": "SG Datalytics SGMPI API",
        "version": "2.0.0",
        "description": "Real Ghana market price data across 6 sectors",
        "endpoints": [
            "GET /api/health",
            "GET /api/stats",
            "GET /api/sectors",
            "GET /api/market-prices?category=&location=&limit=50",
            "GET /api/market-prices/categories",
            "GET /api/market-prices/latest",
            "GET /api/gmpi/latest",
            "GET /api/gmpi",
            "GET /api/gmpi/regional",
            "GET /api/property?location=&limit=50",
            "GET /api/accommodation?type=hotel|airbnb&limit=50",
            "GET /api/economic?indicator=&sector=&limit=50",
            "GET /api/economic/indicators",
            "GET /api/commodities?limit=50",
            "GET /api/financials?limit=50",
            "POST /api/webhook/paystack",
            "GET  /api/download?token=TOKEN",
            "POST /api/test-delivery",
        ]
    })

@app.route('/api/health')
def health():
    results = {}
    total   = 0
    for key in DB:
        try:
            r = query(key, "SELECT COUNT(*) AS n FROM " + {
                'market_prices': 'market_prices',
                'accommodation': 'hotel_prices',
                'property':      'property_prices',
                'economic':      'economic_indicators',
                'commodities':   'commodity_prices',
                'financials':    'stock_prices',
            }[key], one=True)
            results[key] = int(r['n'])
            total += int(r['n'])
        except Exception as e:
            results[key] = f'error: {str(e)[:60]}'
    return jsonify({"status": "ok", "total_records": total, "databases": results})

@app.route('/api/stats')
def stats():
    stats = {}
    try:
        r = query('market_prices', "SELECT COUNT(*) AS n, COUNT(DISTINCT product_category) AS cats, COUNT(DISTINCT COALESCE(NULLIF(location, ''), 'Ghana')) AS locs, MIN(collected_date) AS date_min, MAX(collected_date) AS date_max FROM market_prices", one=True)
        stats['market_prices'] = {'records': int(r['n']), 'categories': int(r['cats']), 'locations': int(r['locs']), 'date_min': str(r['date_min']), 'date_max': str(r['date_max'])}
    except: stats['market_prices'] = {'records': 0}
    try:
        r = query('property', "SELECT COUNT(*) AS n, COUNT(DISTINCT location) AS locs FROM property_prices", one=True)
        stats['property'] = {'records': int(r['n']), 'locations': int(r['locs'])}
    except: stats['property'] = {'records': 0}
    try:
        h = query('accommodation', "SELECT COUNT(*) AS n FROM hotel_prices", one=True)
        a = query('accommodation', "SELECT COUNT(*) AS n FROM airbnb_prices", one=True)
        stats['accommodation'] = {'hotel_records': int(h['n']), 'airbnb_records': int(a['n'])}
    except: stats['accommodation'] = {'records': 0}
    try:
        r = query('economic', "SELECT COUNT(*) AS n, COUNT(DISTINCT indicator_name) AS inds, COUNT(DISTINCT sector) AS secs FROM economic_indicators", one=True)
        stats['economic'] = {'records': int(r['n']), 'indicators': int(r['inds']), 'sectors': int(r['secs'])}
    except: stats['economic'] = {'records': 0}
    try:
        r = query('commodities', "SELECT COUNT(*) AS n FROM commodity_prices", one=True)
        f = query('commodities', "SELECT COUNT(*) AS n FROM fuel_prices", one=True)
        stats['commodities'] = {'commodity_records': int(r['n']), 'fuel_records': int(f['n'])}
    except: stats['commodities'] = {'records': 0}
    try:
        r = query('financials', "SELECT COUNT(*) AS n FROM stock_prices", one=True)
        stats['financials'] = {'records': int(r['n'])}
    except: stats['financials'] = {'records': 0}

    total = sum(v.get('records', 0) + v.get('hotel_records', 0) + v.get('airbnb_records', 0) +
                v.get('commodity_records', 0) + v.get('fuel_records', 0)
                for v in stats.values())
    return jsonify({"total_records": total, "total_sectors": 6, "sectors": stats})

@app.route('/api/sectors')
def sectors():
    return jsonify([
        {"id": 1, "code": "market_prices",  "name": "Market Prices",  "icon": "🛒", "color": "#00957a", "description": "Consumer goods & electronics"},
        {"id": 2, "code": "property",       "name": "Property",       "icon": "🏠", "color": "#2563eb", "description": "Real estate listings"},
        {"id": 3, "code": "accommodation",  "name": "Accommodation",  "icon": "🏨", "color": "#c77c00", "description": "Hotels & Airbnb"},
        {"id": 4, "code": "economic",       "name": "Economic",       "icon": "📈", "color": "#7c3aed", "description": "Macro indicators & FX rates"},
        {"id": 5, "code": "commodities",    "name": "Commodities",    "icon": "⛽", "color": "#dc2626", "description": "Fuel & commodity prices"},
        {"id": 6, "code": "financials",     "name": "Financials",     "icon": "📊", "color": "#059669", "description": "GSE stocks & indices"},
    ])

@app.route('/api/market-prices')
@limiter.limit("60 per minute")
def market_prices():
    category = request.args.get('category')
    location = request.args.get('location')
    limit    = min(int(request.args.get('limit', 50)), 200)
    where, params = [], []
    if category: where.append("product_category ILIKE %s"); params.append(f'%{category}%')
    if location: where.append("location ILIKE %s"); params.append(f'%{location}%')
    sql = "SELECT id, collected_date, week_number, year, product_category, title, price_ghs, location, condition, source FROM market_prices"
    if where: sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY collected_date DESC LIMIT %s"
    params.append(limit)
    rows = query('market_prices', sql, params)
    return jsonify([dict(r) for r in rows])

@app.route('/api/market-prices/categories')
def market_categories():
    rows = query('market_prices', "SELECT DISTINCT product_category, COUNT(*) AS count FROM market_prices GROUP BY product_category ORDER BY count DESC")
    return jsonify([dict(r) for r in rows])

@app.route('/api/market-prices/locations')
def market_locations():
    """
    Diagnostic endpoint — shows how many records have each location value.
    Useful for assessing whether a regional GMPI is feasible.
    """
    rows = query('market_prices', """
        SELECT
          CASE
            WHEN location IS NULL OR TRIM(location) = '' THEN '(no location)'
            ELSE TRIM(location)
          END AS location,
          COUNT(*)                                        AS total_records,
          COUNT(CASE WHEN price_ghs > 0 THEN 1 END)      AS priced_records,
          COUNT(DISTINCT product_category)                AS categories
        FROM market_prices
        GROUP BY 1
        ORDER BY total_records DESC
        LIMIT 100
    """)
    total = query('market_prices', "SELECT COUNT(*) AS n FROM market_prices", one=True)
    blank = query('market_prices',
        "SELECT COUNT(*) AS n FROM market_prices WHERE location IS NULL OR TRIM(location) = ''",
        one=True)
    return jsonify({
        'summary': {
            'total_records':         int(total['n']),
            'records_with_location': int(total['n']) - int(blank['n']),
            'records_without_location': int(blank['n']),
            'pct_with_location': round((int(total['n']) - int(blank['n'])) / int(total['n']) * 100, 1)
        },
        'locations': [dict(r) for r in rows]
    })

@app.route('/api/market-prices/latest')
def market_latest():
    rows = query('market_prices', "SELECT product_category, ROUND(AVG(price_ghs),2) AS avg_price_ghs, COUNT(*) AS listings, MAX(collected_date) AS last_updated FROM market_prices GROUP BY product_category ORDER BY listings DESC")
    return jsonify([dict(r) for r in rows])

@app.route('/api/property')
@limiter.limit("60 per minute")
def property_prices():
    location = request.args.get('location')
    limit    = min(int(request.args.get('limit', 50)), 200)
    where, params = [], []
    if location: where.append("location ILIKE %s"); params.append(f'%{location}%')
    sql = "SELECT * FROM property_prices"
    if where: sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY collected_date DESC LIMIT %s"
    params.append(limit)
    rows = query('property', sql, params)
    return jsonify([dict(r) for r in rows])

@app.route('/api/accommodation')
@limiter.limit("60 per minute")
def accommodation():
    acc_type = request.args.get('type', 'hotel')
    limit    = min(int(request.args.get('limit', 50)), 200)
    table    = 'airbnb_prices' if acc_type == 'airbnb' else 'hotel_prices'
    rows     = query('accommodation', f"SELECT * FROM {table} ORDER BY collected_date DESC LIMIT %s", [limit])
    return jsonify([dict(r) for r in rows])

@app.route('/api/economic')
@limiter.limit("60 per minute")
def economic():
    indicator = request.args.get('indicator')
    sector    = request.args.get('sector')
    limit     = min(int(request.args.get('limit', 50)), 200)
    where, params = [], []
    if indicator: where.append("indicator_name ILIKE %s"); params.append(f'%{indicator}%')
    if sector:    where.append("sector ILIKE %s"); params.append(f'%{sector}%')
    sql = "SELECT id, collected_date, year, month, indicator_code, indicator_name, sector, value, unit, source FROM economic_indicators"
    if where: sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY collected_date DESC LIMIT %s"
    params.append(limit)
    rows = query('economic', sql, params)
    return jsonify([dict(r) for r in rows])

@app.route('/api/economic/indicators')
def economic_indicators():
    rows = query('economic', "SELECT DISTINCT indicator_name, indicator_code, sector, unit, COUNT(*) AS records FROM economic_indicators GROUP BY indicator_name, indicator_code, sector, unit ORDER BY sector, indicator_name")
    return jsonify([dict(r) for r in rows])

@app.route('/api/economic/exchange-rates')
def exchange_rates():
    limit = min(int(request.args.get('limit', 50)), 200)
    rows  = query('economic', "SELECT * FROM exchange_rates ORDER BY collected_date DESC LIMIT %s", [limit])
    return jsonify([dict(r) for r in rows])

@app.route('/api/commodities')
@limiter.limit("60 per minute")
def commodities():
    limit = min(int(request.args.get('limit', 50)), 200)
    rows  = query('commodities', "SELECT * FROM commodity_prices ORDER BY collected_date DESC LIMIT %s", [limit])
    return jsonify([dict(r) for r in rows])

@app.route('/api/fuel')
def fuel():
    limit = min(int(request.args.get('limit', 50)), 200)
    rows  = query('commodities', "SELECT * FROM fuel_prices ORDER BY collected_date DESC LIMIT %s", [limit])
    return jsonify([dict(r) for r in rows])

@app.route('/api/financials/stocks')
def stocks():
    limit = min(int(request.args.get('limit', 50)), 200)
    rows  = query('financials', "SELECT * FROM stock_prices ORDER BY collected_date DESC LIMIT %s", [limit])
    return jsonify([dict(r) for r in rows])

@app.route('/api/financials/indices')
def indices():
    limit = min(int(request.args.get('limit', 50)), 200)
    rows  = query('financials', "SELECT * FROM gse_indices ORDER BY collected_date DESC LIMIT %s", [limit])
    return jsonify([dict(r) for r in rows])

# ═══════════════════════════════════════════════════════════════
# PART 1 — PAYSTACK WEBHOOK
# ═══════════════════════════════════════════════════════════════
@app.route('/api/webhook/paystack', methods=['POST'])
@limiter.limit("30 per minute")
def paystack_webhook():
    secret_key = os.getenv('PAYSTACK_SECRET_KEY', '')
    signature  = request.headers.get('x-paystack-signature', '')
    raw_body   = request.get_data()

    # Verify Paystack HMAC-SHA512 signature
    expected = hmac.new(secret_key.encode('utf-8'), raw_body, hashlib.sha512).hexdigest()
    if not hmac.compare_digest(expected, signature):
        print(f"[WEBHOOK] Invalid signature — possible spoofed request")
        return jsonify({'error': 'Invalid signature'}), 401

    payload = request.get_json(force=True) or {}
    event   = payload.get('event', '')
    print(f"[WEBHOOK] Event received: {event}")

    if event == 'charge.success':
        data      = payload.get('data', {})
        email     = data.get('customer', {}).get('email', '')
        amount    = data.get('amount', 0) / 100   # Paystack sends in pesewas
        meta      = data.get('metadata', {}) or {}

        # Prefer exact product key from metadata (set by frontend triggerPaystack)
        # Fall back to plan-name keyword matching for subscription-style flows
        product_key = meta.get('product', '')
        if product_key in PRODUCT_SECTOR_MAP:
            sector = PRODUCT_SECTOR_MAP[product_key]
        else:
            plan      = data.get('plan', {})
            plan_name = plan.get('name', '') if isinstance(plan, dict) else str(plan)
            sector    = resolve_sector(plan_name)

        if not email:
            print(f"[WEBHOOK] charge.success with no email — skipping")
            return jsonify({'status': 'ok'}), 200

        print(f"[WEBHOOK] ✓ {email} | product='{product_key}' → sector={sector} | GH₵{amount:.2f}")

        if sector == 'bundle':
            # Send a separate download link for each sector in the bundle
            try:
                send_bundle_email(email, BUNDLE_SECTORS)
            except Exception as e:
                print(f"[WEBHOOK] Bundle email failed for {email}: {e}")
        else:
            download_url = generate_download_token(email, sector)
            try:
                send_download_email(email, download_url, sector)
            except Exception as e:
                print(f"[WEBHOOK] Email failed for {email}: {e}")
                # Still return 200 — token was created, email failure is non-fatal

    return jsonify({'status': 'ok'}), 200

# ═══════════════════════════════════════════════════════════════
# PART 3B — GHANA MARKET PRICE INDEX (GMPI)
# ═══════════════════════════════════════════════════════════════

@app.route('/api/gmpi')
@limiter.limit("60 per minute")
def get_gmpi():
    """
    Ghana Market Price Index — weekly composite price index.
    Consumer goods only: Real Estate & Vehicles excluded, price cap GHS 50k.
    Base = all-time median per category (same methodology as /api/gmpi/latest and /api/gmpi/regional).
    """
    # ── Overall weekly GMPI ───────────────────────────────────
    overall_sql = """
    WITH base_medians AS (
      SELECT product_category,
        PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY price_ghs) AS base_median
      FROM market_prices
      WHERE price_ghs BETWEEN 1 AND 50000
        AND product_category NOT IN ('Real Estate', 'Vehicles')
      GROUP BY product_category
      HAVING COUNT(*) >= 20
    ),
    weekly_medians AS (
      SELECT
        week_number, year,
        MIN(collected_date) AS week_date,
        product_category,
        PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY price_ghs) AS wk_median,
        COUNT(*) AS n
      FROM market_prices
      WHERE price_ghs BETWEEN 1 AND 50000
        AND product_category NOT IN ('Real Estate', 'Vehicles')
      GROUP BY week_number, year, product_category
      HAVING COUNT(*) >= 5
    ),
    weekly_cat_idx AS (
      SELECT wm.week_number, wm.year, wm.week_date,
        ROUND((wm.wk_median / bm.base_median * 100)::numeric, 2) AS category_index,
        wm.product_category
      FROM weekly_medians wm
      JOIN base_medians bm USING (product_category)
    ),
    weekly_gmpi AS (
      SELECT week_number, year,
        MIN(week_date) AS week_date,
        ROUND(AVG(category_index)::numeric, 2) AS gmpi,
        COUNT(DISTINCT product_category) AS categories_tracked
      FROM weekly_cat_idx
      GROUP BY week_number, year
      HAVING COUNT(DISTINCT product_category) >= 3
    )
    SELECT
      week_number, year, week_date, gmpi, categories_tracked,
      LAG(gmpi) OVER (ORDER BY year, week_number) AS prev_gmpi,
      ROUND((gmpi - LAG(gmpi) OVER (ORDER BY year, week_number))::numeric, 2) AS gmpi_change,
      ROUND(
        (gmpi - LAG(gmpi) OVER (ORDER BY year, week_number))
        / NULLIF(LAG(gmpi) OVER (ORDER BY year, week_number), 0) * 100
      ::numeric, 2) AS gmpi_change_pct
    FROM weekly_gmpi
    ORDER BY year DESC, week_number DESC
    """

    # ── Per-category weekly index ─────────────────────────────
    category_sql = """
    WITH base_medians AS (
      SELECT product_category,
        PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY price_ghs) AS base_median
      FROM market_prices
      WHERE price_ghs BETWEEN 1 AND 50000
        AND product_category NOT IN ('Real Estate', 'Vehicles')
      GROUP BY product_category
      HAVING COUNT(*) >= 20
    ),
    weekly_medians AS (
      SELECT
        week_number, year,
        MIN(collected_date) AS week_date,
        product_category,
        PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY price_ghs) AS wk_median,
        COUNT(*) AS n
      FROM market_prices
      WHERE price_ghs BETWEEN 1 AND 50000
        AND product_category NOT IN ('Real Estate', 'Vehicles')
      GROUP BY week_number, year, product_category
      HAVING COUNT(*) >= 5
    )
    SELECT
      wm.week_number, wm.year, wm.week_date,
      wm.product_category,
      ROUND((wm.wk_median / bm.base_median * 100)::numeric, 2) AS category_index,
      ROUND(wm.wk_median::numeric, 2) AS median_price_ghs,
      wm.n AS listings_count
    FROM weekly_medians wm
    JOIN base_medians bm USING (product_category)
    ORDER BY wm.year DESC, wm.week_number DESC, wm.product_category
    """

    try:
        overall   = query('market_prices', overall_sql)
        by_cat    = query('market_prices', category_sql)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    return jsonify({
        'index_name':  'Ghana Market Price Index (GMPI)',
        'description': 'Consumer goods only. Real Estate and Vehicles excluded. Base = all-time historical median per category.',
        'base_note':   'GMPI of 110 means prices are 10%% above the historical median.',
        'overall':     overall,
        'by_category': by_cat,
    })


# ── Public GMPI summary (no auth — for website display) ──────
@app.route('/api/gmpi/latest')
@limiter.limit("120 per minute")
def get_gmpi_latest():
    """
    Latest GMPI value + week-on-week change. Safe for public display.
    Methodology: all-time median per category = base (100).
    Consumer goods only — Real Estate & Vehicles excluded, price cap GHS 50k.
    Same methodology as /api/gmpi/regional.
    """
    sql = """
    WITH base_medians AS (
      -- All-time median per category = base (index 100)
      SELECT product_category,
        PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY price_ghs) AS base_median
      FROM market_prices
      WHERE price_ghs BETWEEN 1 AND 50000
        AND product_category NOT IN ('Real Estate', 'Vehicles')
      GROUP BY product_category
      HAVING COUNT(*) >= 20
    ),
    weekly_medians AS (
      -- Per-week, per-category median
      SELECT
        week_number, year,
        MIN(collected_date) AS week_date,
        product_category,
        PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY price_ghs) AS wk_median,
        COUNT(*) AS n
      FROM market_prices
      WHERE price_ghs BETWEEN 1 AND 50000
        AND product_category NOT IN ('Real Estate', 'Vehicles')
      GROUP BY week_number, year, product_category
      HAVING COUNT(*) >= 5
    ),
    weekly_cat_idx AS (
      -- Category index per week = (week median / base median) * 100
      SELECT wm.week_number, wm.year, wm.week_date,
        ROUND((wm.wk_median / bm.base_median * 100)::numeric, 2) AS category_index,
        wm.product_category
      FROM weekly_medians wm
      JOIN base_medians bm USING (product_category)
    ),
    weekly_gmpi AS (
      -- GMPI per week = average of category indices (equal-weighted)
      SELECT week_number, year,
        MIN(week_date) AS week_date,
        ROUND(AVG(category_index)::numeric, 2) AS gmpi,
        COUNT(DISTINCT product_category) AS categories_tracked
      FROM weekly_cat_idx
      GROUP BY week_number, year
      HAVING COUNT(DISTINCT product_category) >= 3
      ORDER BY year DESC, week_number DESC
      LIMIT 2
    )
    SELECT * FROM weekly_gmpi
    ORDER BY year DESC, week_number DESC
    """
    try:
        rows = query('market_prices', sql)
    except Exception as e:
        print(f"[GMPI/latest] SQL error: {e}")
        return jsonify({'error': str(e)}), 500

    if not rows:
        return jsonify({'gmpi': None})

    latest = rows[0]
    prev   = rows[1] if len(rows) > 1 else None

    latest_gmpi = float(latest['gmpi'])
    prev_gmpi   = float(prev['gmpi']) if prev else None
    change      = round(latest_gmpi - prev_gmpi, 2) if prev_gmpi else None
    change_pct  = round((change / prev_gmpi) * 100, 2) if prev_gmpi and change is not None else None

    return jsonify({
        'gmpi':               round(latest_gmpi, 2),
        'week':               latest['week_number'],
        'year':               latest['year'],
        'week_date':          str(latest['week_date']),
        'categories_tracked': latest['categories_tracked'],
        'change':             change,
        'change_pct':         change_pct,
        'label':              f"W{latest['week_number']} {latest['year']}",
    })


# ── Regional GMPI — powers the price map page ────────────────
@app.route('/api/gmpi/regional')
@limiter.limit("60 per minute")
def gmpi_regional():
    """
    Returns GMPI broken down by Ghana region and by Greater Accra neighbourhood.
    Consumer goods only (excludes Real Estate & Vehicles, price cap GHS 50k).
    Base = median across ALL located records per category (100 = national average).
    Both regional and national medians are computed from the same pool of located
    records so the comparison is apples-to-apples.
    A region at 120 is 20%% more expensive than the national average.
    A region at 80 is 20%% cheaper than the national average.
    Powers the interactive price map page.
    """

    # ── Regional GMPI ─────────────────────────────────────────
    # Key fix: national_medians is computed from the SAME pool as regional_medians
    # (only records with an explicit location). This ensures 100 = national average
    # across all located records. Previously national_medians used ALL records
    # (including unlabelled listings) which skewed the baseline downward and
    # produced inflated values like Greater Accra = 299.8.
    regional_sql = """
    WITH located_data AS (
      SELECT product_category, price_ghs, location
      FROM market_prices
      WHERE price_ghs BETWEEN 1 AND 50000
        AND product_category NOT IN ('Real Estate', 'Vehicles')
        AND TRIM(location) != ''
        AND location NOT ILIKE 'Nationwide%%'
    ),
    national_medians AS (
      -- Median across ALL located records = the benchmark (100)
      SELECT product_category,
        PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY price_ghs) AS national_median
      FROM located_data
      GROUP BY product_category
      HAVING COUNT(*) >= 20
    ),
    regional_medians AS (
      SELECT
        CASE
          WHEN location ILIKE 'Greater Accra%%' THEN 'Greater Accra'
          WHEN location ILIKE 'Ashanti%%'       THEN 'Ashanti'
          WHEN location ILIKE 'Western%%'       THEN 'Western Region'
          WHEN location ILIKE 'Central%%'       THEN 'Central Region'
          WHEN location ILIKE 'Eastern%%'       THEN 'Eastern Region'
          WHEN location ILIKE 'Northern%%'      THEN 'Northern Region'
          WHEN location ILIKE 'Brong%%'         THEN 'Brong Ahafo'
        END AS region,
        product_category,
        PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY price_ghs) AS region_median,
        COUNT(*) AS n
      FROM located_data
      GROUP BY 1, 2
      HAVING COUNT(*) >= 10
    ),
    regional_cat_idx AS (
      SELECT rm.region, rm.product_category,
        ROUND((rm.region_median / nm.national_median * 100)::numeric, 2) AS category_index,
        rm.n
      FROM regional_medians rm
      JOIN national_medians nm USING (product_category)
      WHERE rm.region IS NOT NULL
    )
    SELECT region AS name,
      ROUND(AVG(category_index)::numeric, 2) AS gmpi,
      SUM(n)::int AS records,
      COUNT(DISTINCT product_category)::int AS categories
    FROM regional_cat_idx
    GROUP BY region
    HAVING COUNT(DISTINCT product_category) >= 3
    ORDER BY records DESC
    """

    # ── Neighbourhood GMPI (Greater Accra only) ───────────────
    # Same fix: national benchmark uses only located records so 100 = national avg
    nbhd_sql = """
    WITH located_data AS (
      SELECT product_category, price_ghs, location
      FROM market_prices
      WHERE price_ghs BETWEEN 1 AND 50000
        AND product_category NOT IN ('Real Estate', 'Vehicles')
        AND TRIM(location) != ''
        AND location NOT ILIKE 'Nationwide%%'
    ),
    national_medians AS (
      SELECT product_category,
        PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY price_ghs) AS national_median
      FROM located_data
      GROUP BY product_category
      HAVING COUNT(*) >= 20
    ),
    nbhd_medians AS (
      SELECT
        TRIM(SPLIT_PART(location, ', ', 2)) AS neighbourhood,
        product_category,
        PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY price_ghs) AS nbhd_median,
        COUNT(*) AS n
      FROM located_data
      WHERE location ILIKE 'Greater Accra, %%'
      GROUP BY 1, 2
      HAVING COUNT(*) >= 5
    ),
    nbhd_cat_idx AS (
      SELECT nm.neighbourhood, nm.product_category,
        ROUND((nm.nbhd_median / nat.national_median * 100)::numeric, 2) AS category_index,
        nm.n
      FROM nbhd_medians nm
      JOIN national_medians nat USING (product_category)
    )
    SELECT neighbourhood AS name,
      ROUND(AVG(category_index)::numeric, 2) AS gmpi,
      SUM(n)::int AS records,
      COUNT(DISTINCT product_category)::int AS categories
    FROM nbhd_cat_idx
    GROUP BY neighbourhood
    HAVING COUNT(DISTINCT product_category) >= 3
    ORDER BY records DESC
    """

    try:
        regional_rows = query('market_prices', regional_sql)
        nbhd_rows     = query('market_prices', nbhd_sql)
        return jsonify({
            'regions':        [dict(r) for r in regional_rows],
            'neighbourhoods': [dict(r) for r in nbhd_rows],
        })
    except Exception as e:
        print(f"[GMPI/regional] {e}")
        return jsonify({'error': str(e)}), 500

# ═══════════════════════════════════════════════════════════════
# PART 4 — DOWNLOAD ENDPOINT
# ═══════════════════════════════════════════════════════════════
def _lookup_token(token):
    """Fetch token record from Neon. Returns (record, error_response)."""
    try:
        conn = _get_token_conn()
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM download_tokens WHERE token = %s", (token,))
            record = cur.fetchone()
        conn.close()
        return record, None
    except Exception as e:
        print(f"[DOWNLOAD] Token DB lookup failed: {e}")
        return None, (jsonify({'error': 'Service error — please try again shortly.'}), 503)

def _validate_token(record):
    """Check token is valid, not expired, not used. Returns error string or None."""
    if not record:
        return 'Invalid or expired token. Please contact support.'
    now = datetime.now(timezone.utc)
    if now > record['expires_at']:
        return 'This download link has expired (valid for 24 hrs). Please purchase again.'
    if record['used']:
        return 'This link has already been used. Check your email for your file, or contact support.'
    return None


@app.route('/api/download', methods=['GET'])
@limiter.limit("30 per minute")
def download_landing():
    """
    GET /api/download?token=...
    Returns an HTML landing page with a single Download button.
    Email pre-fetchers (Gmail, Outlook etc.) hit this URL automatically to
    scan for malware — they read HTML but won't POST a form, so the token
    stays valid until the real user clicks.
    """
    token = request.args.get('token', '').strip()
    if not token:
        return "<h2>Invalid link — no token provided.</h2>", 400

    record, err = _lookup_token(token)
    if err:
        return err

    error = _validate_token(record)
    sector_label = SECTOR_LABELS.get(record['sector'] if record else '', 'Dataset') if record else 'Dataset'

    if error:
        html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
        <title>Download — SG Datalytics</title>
        <style>body{{font-family:Arial,sans-serif;max-width:560px;margin:80px auto;padding:0 24px;color:#1a1a2e;}}
        h2{{color:#c0392b;}}.logo{{font-weight:700;font-size:1.2rem;color:#1a1a2e;margin-bottom:2rem;display:block;}}
        </style></head><body>
        <span class="logo">SG Datalytics</span>
        <h2>Link unavailable</h2>
        <p>{error}</p>
        <p>Need help? Reply to your purchase confirmation email.</p>
        </body></html>"""
        return html, 410

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
    <title>Download — SG Datalytics</title>
    <style>
      *{{box-sizing:border-box;margin:0;padding:0;}}
      body{{font-family:Arial,sans-serif;background:#f7f8fc;min-height:100vh;display:flex;align-items:center;justify-content:center;}}
      .card{{background:#fff;border-radius:12px;padding:48px 40px;max-width:480px;width:100%;box-shadow:0 2px 16px rgba(0,0,0,0.08);text-align:center;}}
      .logo{{font-weight:700;font-size:1.1rem;color:#1a1a2e;margin-bottom:2rem;display:block;letter-spacing:.5px;}}
      .icon{{font-size:3rem;margin-bottom:1rem;}}
      h1{{font-size:1.4rem;font-weight:700;color:#1a1a2e;margin-bottom:.5rem;}}
      p{{color:#555;font-size:.95rem;line-height:1.6;margin-bottom:1.5rem;}}
      form{{margin-top:1rem;}}
      button{{background:#1a1a2e;color:#fff;border:none;padding:14px 32px;border-radius:8px;font-size:1rem;font-weight:600;cursor:pointer;width:100%;transition:background .2s;}}
      button:hover{{background:#2d2d5e;}}
      .note{{font-size:.8rem;color:#999;margin-top:1rem;}}
    </style></head><body>
    <div class="card">
      <span class="logo">SG Datalytics</span>
      <div class="icon">📦</div>
      <h1>Your data is ready</h1>
      <p><strong>{sector_label}</strong><br>Click the button below to download your CSV file. This link is valid for 24 hours.</p>
      <form method="POST" action="/api/download?token={token}">
        <button type="submit">⬇ Download your file</button>
      </form>
      <p class="note">One-time download · CSV format · Secure link</p>
    </div>
    </body></html>"""
    return html


@app.route('/api/download', methods=['POST'])
@limiter.limit("10 per minute")
def download_data():
    """
    POST /api/download?token=...
    Triggered by the user clicking the Download button on the landing page.
    Marks token as used and streams the CSV file.
    """
    token = request.args.get('token', '').strip()
    if not token:
        return jsonify({'error': 'Token required'}), 400

    record, err = _lookup_token(token)
    if err:
        return err

    error = _validate_token(record)
    if error:
        return f"<h2>{error}</h2>", 410

    # Mark as used in Neon immediately before streaming
    try:
        conn = _get_token_conn()
        with conn:
            with conn.cursor() as cur:
                cur.execute("UPDATE download_tokens SET used = TRUE WHERE token = %s", (token,))
        conn.close()
    except Exception as e:
        print(f"[DOWNLOAD] Failed to mark token used: {e}")

    sector   = record['sector']
    email    = record['email']
    now      = datetime.now(timezone.utc)
    date_str = now.strftime('%Y-%m-%d')
    filename = f"sgdatalytics_{sector}_{date_str}.xlsx"
    queries  = SECTOR_QUERIES.get(sector, [])

    # Build Excel workbook — one sheet per table
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill(fill_type='solid', fgColor='1A1A2E')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Friendly sheet name map (table_label → short Excel tab name, max 31 chars)
    SHEET_NAMES = {
        'market_prices_category_summary': 'Category Summary',
        'market_prices_product_detail':   'Product Detail',
        'hotel_prices_aggregated':        'Hotels',
        'airbnb_prices_aggregated':       'Airbnb',
        'property_prices_aggregated':     'Property',
        'economic_indicators':            'Economic Indicators',
        'exchange_rates':                 'Exchange Rates',
        'gse_indices':                    'GSE Indices',
        'stock_prices':                   'Stock Prices',
        'commodity_prices':               'Commodity Prices',
        'fuel_prices':                    'Fuel Prices',
    }

    for entry in queries:
        db_key, sql, table_label = entry if len(entry) == 3 else (*entry, '')
        try:
            rows = query(db_key, sql)
        except Exception as e:
            print(f"[DOWNLOAD] DB error for {db_key}: {e}")
            continue

        if not rows:
            continue

        sheet_name = SHEET_NAMES.get(table_label, table_label[:31])
        ws = wb.create_sheet(title=sheet_name)

        # Header row
        headers = list(rows[0].keys())
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace('_', ' ').title())
            cell.font  = header_font
            cell.fill  = header_fill
            cell.alignment = header_align

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, key in enumerate(headers, start=1):
                val = row.get(key)
                # Keep numbers as numbers so Excel can sort/sum them
                if val is not None:
                    try:
                        val = float(val) if '.' in str(val) else int(val)
                    except (ValueError, TypeError):
                        val = str(val)
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto-fit column widths (capped at 50)
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        # Freeze top row
        ws.freeze_panes = 'A2'

    print(f"[DOWNLOAD] {email} downloaded '{sector}' at {now.isoformat()}")

    # Write workbook to bytes buffer and serve
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

# ═══════════════════════════════════════════════════════════════
# TEST ENDPOINT — manual trigger without Paystack payment
# ═══════════════════════════════════════════════════════════════
@app.route('/api/test-delivery', methods=['POST'])
@limiter.limit("5 per minute")
def test_delivery():
    """
    Test the full delivery pipeline without a real Paystack payment.
    Body: { "email": "test@example.com", "sector": "market_prices" }
    """
    data   = request.get_json() or {}
    email  = data.get('email', '').strip()
    sector = data.get('sector', 'market_prices').strip()

    if not email:
        return jsonify({'error': 'email is required'}), 400

    valid_sectors = list(SECTOR_QUERIES.keys())
    if sector not in valid_sectors:
        return jsonify({'error': f'Invalid sector. Choose from: {valid_sectors}'}), 400

    download_url = generate_download_token(email, sector, expires_hours=24)

    try:
        send_download_email(email, download_url, sector)
        return jsonify({
            'status':       'ok',
            'message':      f'Email sent to {email}',
            'sector':       sector,
            'download_url': download_url,
        })
    except Exception as e:
        return jsonify({
            'status':       'email_failed',
            'message':      str(e),
            'download_url': download_url,  # still return URL so you can test the download manually
        }), 500

# ── ERROR HANDLERS ───────────────────────────────────────────
@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "Endpoint not found", "available": "/"}), 404

@app.errorhandler(429)
def rate_limited(e):
    return jsonify({"error": "Too many requests — please slow down"}), 429

@app.errorhandler(500)
def server_error(e):
    return jsonify({"error": "Internal server error", "message": str(e)}), 500

# ── START ─────────────────────────────────────────────────────
if __name__ == '__main__':
    port = int(os.getenv('PORT', os.getenv('FLASK_PORT', 5050)))
    print(f"\n  SG Datalytics SGMPI API v2 → http://0.0.0.0:{port}")
    print(f"  Connected databases: {[k for k,v in DB.items() if v]}\n")
    app.run(host='0.0.0.0', port=port, debug=os.getenv('FLASK_ENV') != 'production')
